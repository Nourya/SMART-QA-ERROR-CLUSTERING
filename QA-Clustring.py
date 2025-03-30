import streamlit as st
import xml.etree.ElementTree as ET
import re
import pandas as pd
import io
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Analyse Robot Framework", layout="wide")
st.title("🤖 Analyse intelligente des erreurs Robot Framework")

uploaded_file = st.file_uploader("📂 Téléverse ton fichier `output.xml`", type="xml")

# Extraction des tests échoués
def extract_failed_tests(root):
    failed_tests = []
    for test in root.findall(".//test"):
        test_name = test.attrib.get("name", "Test inconnu")
        status = test.find("status")
        if status is not None and status.attrib.get("status") == "FAIL":
            errors = []
            for msg in test.findall(".//kw/status[@status='FAIL']"):
                if msg.text and msg.text.strip():
                    errors.append(msg.text.strip())
            if errors:
                failed_tests.append({
                    "test_name": test_name,
                    "errors": errors
                })
    return failed_tests

# Nettoyage et simplification
def clean_text(text):
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def simplify_message(msg, max_len=300):
    msg = re.sub(r'Stacktrace:.*', '', msg, flags=re.DOTALL)
    msg = re.sub(r'\s+', ' ', msg).strip()
    return msg[:max_len] + "..." if len(msg) > max_len else msg

# Analyse type GPT locale
def suggest_fix(error_msg):
    msg = error_msg.lower()
    if "not visible" in msg:
        return ("L'élément est peut-être masqué ou trop lent à charger.",
                "Ajoute un `Wait Until Element Is Visible` plus long ou vérifie la condition d'affichage.")
    elif "not found" in msg:
        return ("Le sélecteur ne correspond plus à un élément valide.",
                "Vérifie que l'ID ou le locator est toujours correct dans la page HTML.")
    elif "click intercepted" in msg:
        return ("Un élément bloque le clic (ex: pop-up, spinner...)", 
                "Attends la disparition du blocage avec `Wait Until Element Is Not Visible`.")
    elif "timeout" in msg:
        return ("Le test attend une condition qui ne se réalise pas à temps.",
                "Augmente le timeout ou vérifie la logique conditionnelle.")
    else:
        return ("Erreur générique", "Vérifie les logs complets ou isole le test concerné.")

# Traitement du fichier
if uploaded_file:
    tree = ET.parse(uploaded_file)
    root = tree.getroot()

    all_tests = root.findall(".//test")
    total_tests = len(all_tests)
    failed_tests_data = extract_failed_tests(root)
    failed_count = len(failed_tests_data)
    passed_count = total_tests - failed_count

    st.markdown(f"✅ **Total tests : {total_tests}** | 🟢 Réussis : {passed_count} | 🔴 Échoués : {failed_count}")

    if failed_tests_data:
        st.subheader("📌 Détails des tests échoués")
        full_errors = []

        for test in failed_tests_data:
            st.markdown(f"### ❌ {test['test_name']}")
            for idx, error in enumerate(test['errors'], 1):
                simplified = simplify_message(error)
                cause, fix = suggest_fix(error)
                st.markdown(f"- **Erreur {idx}** : {simplified}")
                st.markdown(f"  - 💡 *Cause probable* : _{cause}_")
                st.markdown(f"  - 🛠️ *Suggestion* : _{fix}_")

                full_errors.append({
                    "Test": test["test_name"],
                    "Erreur complète": error,
                    "Erreur simplifiée": simplified,
                    "Cause probable": cause,
                    "Suggestion de correction": fix
                })

        # Clustering intelligent
        st.subheader("🧠 Regroupement des erreurs similaires (clustering)")
        cleaned = [clean_text(e["Erreur complète"]) for e in full_errors]
        vectorizer = TfidfVectorizer()
        X = vectorizer.fit_transform(cleaned)
        n_clusters = min(3, len(full_errors))
        kmeans = KMeans(n_clusters=n_clusters, random_state=42)
        kmeans.fit(X)

        for i, label in enumerate(kmeans.labels_):
            full_errors[i]["Groupe"] = f"Groupe {label+1}"

        # Affichage regroupé
        grouped = {}
        for e in full_errors:
            grouped.setdefault(e["Groupe"], []).append(e)

        for group in sorted(grouped.keys()):
            messages = grouped[group]
            st.markdown(f"### 🧠 {group} ({len(messages)} erreurs)")
            for i, msg in enumerate(messages, 1):
                st.markdown(f"**{i}.** *{msg['Erreur simplifiée']}*")
        import plotly.express as px

        # 🔢 Données de graphe : compter les erreurs par groupe
        group_counts = pd.Series([e["Groupe"] for e in full_errors]).value_counts().reset_index()
        group_counts.columns = ["Groupe", "Nombre des erreurs"]

        st.subheader("📈 Répartition visuelle des groupes d'erreurs")
        fig = px.pie(group_counts, names='Groupe', values='Nombre des erreurs',
                  title="Répartition des erreurs par groupe",
                  color_discrete_sequence=px.colors.qualitative.Set3,
                  hole=0.4)

        fig.update_traces(textinfo='percent+label', textfont_size=14)
        st.plotly_chart(fig, use_container_width=True)



        # Génération Excel stylé
        st.subheader("📥 Export Excel structuré & stylé")
        df = pd.DataFrame(full_errors)

        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Erreurs groupées"

        headers = list(df.columns)
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        group_colors = ["DCE6F1", "E2EFDA", "FCE4D6", "F9D5E5"]

        for col_num, col_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 30

        for i, row in df.iterrows():
            group_num = int(re.search(r'\d+', row["Groupe"]).group())
            fill_color = group_colors[(group_num - 1) % len(group_colors)]
            for j, val in enumerate(row, 1):
                cell = ws.cell(row=i+2, column=j, value=val)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        wb.save(output)
        output.seek(0)

        st.download_button(
            label="📊 Télécharger Excel structuré (.xlsx)",
            data=output,
            file_name="erreurs_robot_framework.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Aucune erreur détectée dans le fichier. Bravo ! 🎉")
